import openpyxl

def getRowCount(file, sheetName):
   workbook = openpyxl.load_workbook(file)
   sheet = workbook.get_sheet_by_name(sheetName)
   return(sheet.max_row)


def getColumnCount(file, sheetName):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook.get_sheet_by_name(sheetName)
    return(sheet.max_column)


def ReadData(file: object, sheetName: object, rownum: object, columnno: object) -> object:
    workbook = openpyxl.load_workbook(file)
    sheet = workbook.get_sheet_by_name(sheetName)
    return sheet.cell(row=rownum, column=columnno).value


def WriteData(file: object, sheetName: object, rownum: object, columnno: object, data: object) -> object:
    workbook = openpyxl.load_workbook(file)
    sheet = workbook.get_sheet_by_name(sheetName)
    sheet.cell(row=rownum, column=columnno).value = data
    workbook.save(file)